"""Generate user-test materials, attorney attachment, and results spreadsheet.

Run from any directory; outputs go to the script's own folder.
"""
from __future__ import annotations
import os
from datetime import date
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor, black, white
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle,
    KeepTogether, ListFlowable, ListItem, HRFlowable
)
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.workbook.defined_name import DefinedName

HERE = os.path.dirname(os.path.abspath(__file__))

GREEN_DARK   = HexColor("#0a4f3c")
GREEN_LIGHT  = HexColor("#d8ebe1")
ORANGE       = HexColor("#e9763a")
GRAY_900     = HexColor("#1a1917")
GRAY_700     = HexColor("#3a3835")
GRAY_500     = HexColor("#6b6863")
GRAY_200     = HexColor("#dedcd9")
GRAY_100     = HexColor("#ededeb")
GRAY_50      = HexColor("#f7f7f6")
RED          = HexColor("#c53030")


def styles():
    s = getSampleStyleSheet()
    base_font = "Helvetica"
    bold_font = "Helvetica-Bold"
    out = {
        "h1": ParagraphStyle("h1", parent=s["Heading1"],
                             fontName=bold_font, fontSize=22, leading=26,
                             textColor=GRAY_900, spaceAfter=10, spaceBefore=0),
        "h2": ParagraphStyle("h2", parent=s["Heading2"],
                             fontName=bold_font, fontSize=14, leading=18,
                             textColor=GREEN_DARK, spaceAfter=6, spaceBefore=18),
        "h3": ParagraphStyle("h3", parent=s["Heading3"],
                             fontName=bold_font, fontSize=11, leading=14,
                             textColor=GRAY_900, spaceAfter=4, spaceBefore=10),
        "body": ParagraphStyle("body", parent=s["Normal"],
                               fontName=base_font, fontSize=10, leading=14,
                               textColor=GRAY_900, spaceAfter=6,
                               alignment=TA_LEFT),
        "muted": ParagraphStyle("muted", parent=s["Normal"],
                                fontName=base_font, fontSize=9, leading=12,
                                textColor=GRAY_500, spaceAfter=4),
        "lead": ParagraphStyle("lead", parent=s["Normal"],
                               fontName=base_font, fontSize=11, leading=15,
                               textColor=GRAY_700, spaceAfter=10),
        "quote": ParagraphStyle("quote", parent=s["Normal"],
                                fontName="Helvetica-Oblique", fontSize=10,
                                leading=14, textColor=GRAY_700,
                                leftIndent=14, rightIndent=14,
                                spaceAfter=8, spaceBefore=4,
                                borderPadding=(8, 10, 8, 10),
                                backColor=GREEN_LIGHT),
        "callout": ParagraphStyle("callout", parent=s["Normal"],
                                  fontName=bold_font, fontSize=10,
                                  leading=14, textColor=ORANGE,
                                  spaceAfter=6, spaceBefore=4),
        "tiny": ParagraphStyle("tiny", parent=s["Normal"],
                               fontName=base_font, fontSize=8, leading=10,
                               textColor=GRAY_500),
        "bullet": ParagraphStyle("bullet", parent=s["Normal"],
                                 fontName=base_font, fontSize=10, leading=14,
                                 textColor=GRAY_900, spaceAfter=2,
                                 leftIndent=14, bulletIndent=4),
        "footer": ParagraphStyle("footer", parent=s["Normal"],
                                 fontName=base_font, fontSize=8, leading=10,
                                 textColor=GRAY_500, alignment=TA_CENTER),
    }
    return out


def header_footer(canvas, doc, title, subtitle=""):
    canvas.saveState()
    # Top header
    canvas.setFillColor(GREEN_DARK)
    canvas.rect(0, A4[1] - 14 * mm, A4[0], 14 * mm, fill=1, stroke=0)
    canvas.setFillColor(white)
    canvas.setFont("Helvetica-Bold", 9)
    canvas.drawString(20 * mm, A4[1] - 9 * mm, "MittBygg")
    canvas.setFont("Helvetica", 9)
    canvas.drawRightString(A4[0] - 20 * mm, A4[1] - 9 * mm, title)
    # Footer
    canvas.setFillColor(GRAY_500)
    canvas.setFont("Helvetica", 8)
    canvas.drawString(20 * mm, 12 * mm, subtitle)
    canvas.drawRightString(A4[0] - 20 * mm, 12 * mm, f"Side {doc.page}")
    canvas.restoreState()


# ============================================================================
# 1. USER-TEST SCRIPT PDF
# ============================================================================
def build_test_script():
    out = os.path.join(HERE, "brukertest_skript.pdf")
    s = styles()
    doc = SimpleDocTemplate(
        out, pagesize=A4,
        leftMargin=20 * mm, rightMargin=20 * mm,
        topMargin=22 * mm, bottomMargin=18 * mm,
        title="MittBygg — Brukertest-skript",
        author="MittBygg",
    )
    story = []

    # Title
    story.append(Paragraph("Brukertest-skript", s["h1"]))
    story.append(Paragraph(
        "30-minutters-test som avdekker hvor folk faller av i MittBygg-flyten. "
        "Skript til moderator. Skriv ut og les fra.", s["lead"]))
    story.append(HRFlowable(width="100%", color=GRAY_200, thickness=0.5,
                            spaceAfter=10))

    # Meta-table
    meta = [
        ["Tester (navn):", "_____________________________"],
        ["Profil-nr:", "1   2   3   4   5   (sirkle)"],
        ["Dato:", "_____________________________"],
        ["Sted / kanal:", "Hjemme  /  Kontor  /  Video"],
        ["Moderator:", "_____________________________"],
        ["Varighet:", "_______ min  (mål: 30)"],
    ]
    t = Table(meta, colWidths=[40 * mm, 90 * mm])
    t.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), "Helvetica", 10),
        ("FONT", (0, 0), (0, -1), "Helvetica-Bold", 10),
        ("TEXTCOLOR", (0, 0), (0, -1), GRAY_700),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, GRAY_200),
    ]))
    story.append(t)
    story.append(Spacer(1, 8))

    # Intro
    story.append(Paragraph("Før du starter (2 min)", s["h2"]))
    story.append(Paragraph(
        "<b>Sett opp</b>: telefon eller PC med <b>http://127.0.0.1:8765/"
        "index.html</b> (eller Netlify-lenken) klar i nettleser. "
        "Lyd-opptak hvis tester samtykker. Notatblokk for sitater. "
        "Ikke vis dem startsiden før de er klare.",
        s["body"]))
    story.append(Paragraph(
        "<b>Mål for testen</b>: avdekk hvor de blir forvirret, hvor de "
        "mister tillit, og hva de ville betalt. Ikke selg appen. "
        "Bli forvirret SAMMEN med dem hvis de spør &quot;hva gjør jeg her?&quot;.",
        s["body"]))

    # Opening script
    story.append(Paragraph("Åpningsmanus (les ordrett)", s["h2"]))
    story.append(Paragraph(
        "&quot;Hei. Jeg jobber med en app som skal hjelpe vanlige folk med "
        "byggesøknader. Den er ikke ferdig — det er derfor jeg vil ha din hjelp."
        "<br/><br/>"
        "Det viktigste du kan gjøre er å bli forvirret og si det høyt. "
        "Jeg lærer mer av at du ikke skjønner noe enn av at du syns alt er fint."
        "<br/><br/>"
        "Jeg sitter ved siden av deg, men jeg kommer ikke til å hjelpe. "
        "Hvis du spør meg &quot;hva trykker jeg på?&quot; så svarer jeg "
        "&quot;hva tror du selv?&quot;. Det er vanskelig, men det er "
        "kjernen i hva jeg trenger fra deg. "
        "Klart?&quot;",
        s["quote"]))

    story.append(Paragraph("Sett scenen", s["h3"]))
    story.append(Paragraph(
        "&quot;Forestill deg at du har en uinnredet kjeller du vurderer å "
        "gjøre om til soverom. Du har akkurat åpnet denne appen for første "
        "gang. Vis meg hva du gjør. Snakk høyt mens du klikker.&quot;",
        s["quote"]))

    story.append(PageBreak())

    # Tasks
    story.append(Paragraph("Oppgaver — observer, ikke hjelp", s["h2"]))
    story.append(Paragraph(
        "Følg rekkefølgen. Etter hver oppgave: noter <b>tid brukt</b>, "
        "<b>nølinger</b>, og <b>sitater</b>.", s["muted"]))

    tasks = [
        ("1. Logg inn", "Be tester logge inn med &quot;BankID&quot;. "
         "Observer: forstår de mock-knappen? Klikker de &quot;Prøv demo&quot; "
         "isteden?"),
        ("2. Søk din egen adresse", "Be dem prøve å søke sin <i>ekte</i> "
         "adresse. Den finnes ikke i demo-data. <b>Observer reaksjonen</b> "
         "— blir de irritert? Skjønner de at det er demo? Mister de tillit?"),
        ("3. Velg Solhellinga 24", "&quot;La oss late som dette er huset "
         "ditt.&quot; Trykker de på den i lista? Forstår de at den er "
         "valgbar?"),
        ("4. Bli kjent med eiendommen", "Stilltiende. &quot;Hva er nytt "
         "for deg her?&quot; Skjønner de gnr/bnr, byggegrenser, "
         "tegninger på arkiv? Klikker de på &quot;Se alle (N)&quot; "
         "for tegningsarkivet?"),
        ("5. Start tiltak", "&quot;Du vurderer å gjøre om kjelleren til "
         "soverom. Vis meg hvordan du starter.&quot; Finner de "
         "&quot;Hva vil du gjøre?&quot;-knappen? Velger de riktig tiltak?"),
        ("6. Gjennom kjeller-wizarden",
         "4 steg: rom → ny bruk → tilstand → bekreft. "
         "Observer hvert steg: skjønner de spørsmålene? Hvilke "
         "begreper er forvirrende? "
         "(Ofte forvirrende: &quot;balansert ventilasjon&quot;, &quot;Bq/m³&quot;, "
         "&quot;hybel vs soverom&quot;.)"),
        ("7. Resultat-siden", "&quot;Hva ville du gjort nå?&quot; "
         "Stoler de på resultatet? Leser de PBL-paragrafene? "
         "Forstår de kostnadsestimatet? "
         "Trykker de på &quot;Generer søknadspakke&quot;?"),
        ("8. Søknadspakken", "Bla gjennom tegninger og dokumenter. "
         "Spør: &quot;Hva tenker du om dette?&quot; Sjekk om de "
         "skjønner forskjellen mellom standard- og premium-pakke."),
    ]
    for title, desc in tasks:
        story.append(Paragraph(f"<b>{title}</b>", s["h3"]))
        story.append(Paragraph(desc, s["body"]))
        story.append(Paragraph(
            "<i>Tid: ____  Nølinger: ____  Falt av: Ja / Nei</i>",
            s["muted"]))
        story.append(Spacer(1, 4))

    story.append(PageBreak())

    # Debrief
    story.append(Paragraph("Etter testen — 7 spørsmål (5 min)", s["h2"]))
    story.append(Paragraph(
        "Disse er handlingsorienterte. Ikke spør &quot;hva syns du om "
        "appen?&quot; — folk er hyggelige og lyver. Spør konkret.",
        s["muted"]))

    debrief = [
        ("Hvilken del var mest forvirrende?",
         "De mest ærlige svarene kommer her. Følg opp med &quot;hva forventet "
         "du å skje?&quot;."),
        ("Hva trodde du skulle skje da du trykket [...]?",
         "Spør om HVERT sted hvor de nølte. Ikke skip noen."),
        ("På en skala 1–10: hvor mye stoler du på regelsjekken?",
         "Følg opp: &quot;Hva ville fått deg til å stole mer? Arkitekt-signatur? "
         "Kommunens stempel? Prøveperiode med pengene-tilbake?&quot;"),
        ("Hva ville du betalt for søknadspakken?",
         "<b>Ikke gi alternativer.</b> La dem nevne tall først. Skriv ned. "
         "Etter de har nevnt tall: &quot;hva med 4 900 kr? 12 900 kr "
         "med arkitekt-KS?&quot;"),
        ("Hvis du virkelig hadde et kjellerprosjekt — hadde du brukt "
         "dette eller ringt arkitekt?",
         "Følg opp: &quot;Hvorfor?&quot; Hvis de svarer arkitekt: &quot;Hva "
         "ville fått deg til å bytte til appen?&quot;"),
        ("Hva manglet?",
         "Åpent. Hvis de drar på det: &quot;Hva med på siste skjerm? "
         "Hva med på første?&quot;"),
        ("Hvem du kjenner ville hatt nytte av dette?",
         "Hvis de nevner 2+ navn = sterkt signal. "
         "Spør: &quot;Kan jeg sende lenken til dem og si du anbefalte?&quot;"),
    ]
    for i, (q, hint) in enumerate(debrief, 1):
        story.append(Paragraph(f"<b>{i}. {q}</b>", s["h3"]))
        story.append(Paragraph(f"<i>Hint:</i> {hint}", s["muted"]))
        story.append(Paragraph(
            "Svar: ______________________________________________________"
            "______________________________________________________"
            "______________________________________________________",
            s["body"]))
        story.append(Spacer(1, 4))

    # Closing
    story.append(Paragraph("Avslutning", s["h2"]))
    story.append(Paragraph(
        "&quot;Tusen takk. Det du har sagt nå er mer verdt enn alt jeg "
        "kunne lest om byggesøknad-prosessen. Jeg sender deg lenken til "
        "den ferdige versjonen når den er klar — om du vil prøve den med "
        "en ekte sak.&quot;",
        s["quote"]))

    story.append(Paragraph("Etter alle 5 testene", s["h2"]))
    story.append(Paragraph(
        "<b>Se etter mønstre, ikke gjennomsnitt.</b><br/>"
        "• Hvis 4/5 falt av samme sted → fix den ÉNE tingen før Stage 1.<br/>"
        "• Hvis tillit &lt; 6/10 hos 3+ → endre Stage 1: arkitekt-KS som "
        "default, ikke premium.<br/>"
        "• Hvis ingen ville betalt &gt; 2 000 kr → endre forretningsmodell.<br/>"
        "• Hvis 3+ nevnte samme manglende funksjon → den må inn i Sprint 1.",
        s["body"]))

    # Render
    def cb(canv, doc_):
        header_footer(canv, doc_,
                      title="Brukertest-skript",
                      subtitle=f"v1.0 · {date.today().strftime('%d.%m.%Y')}")
    doc.build(story, onFirstPage=cb, onLaterPages=cb)
    print(f"Wrote {out}")


# ============================================================================
# 2. RESULTS SPREADSHEET
# ============================================================================
def build_spreadsheet():
    out = os.path.join(HERE, "brukertest_resultater.xlsx")
    wb = Workbook()

    # Styles
    header_fill = PatternFill("solid", fgColor="0a4f3c")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="d8ebe1")
    section_font = Font(name="Calibri", size=10, bold=True, color="0a4f3c")
    instruction_font = Font(name="Calibri", size=9, italic=True, color="6b6863")
    thin = Side(border_style="thin", color="dedcd9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ----- Sheet 1: Recording sheet -----
    ws = wb.active
    ws.title = "Resultater"
    ws.sheet_view.zoomScale = 110

    # Top header
    ws["A1"] = "MittBygg — Brukertest-resultater"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="0a4f3c")
    ws.merge_cells("A1:M1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = (f"Skjemaet fylles ut PR. TESTER. "
                f"Versjon: v1.0 · {date.today().strftime('%d.%m.%Y')} · "
                f"En rad pr tester. Bruk Sammendrag-arket for mønster-analyse.")
    ws["A2"].font = instruction_font
    ws.merge_cells("A2:M2")
    ws.row_dimensions[2].height = 20

    # Section: Identifikasjon
    ws["A4"] = "IDENTIFIKASJON"
    ws["A4"].fill = section_fill
    ws["A4"].font = section_font
    ws.merge_cells("A4:D4")

    ws["E4"] = "OBSERVASJONER"
    ws["E4"].fill = section_fill
    ws["E4"].font = section_font
    ws.merge_cells("E4:I4")

    ws["J4"] = "FORRETNING"
    ws["J4"].fill = section_fill
    ws["J4"].font = section_font
    ws.merge_cells("J4:M4")

    headers = [
        ("A", "Tester #",            8),
        ("B", "Navn",               18),
        ("C", "Profil",             18),
        ("D", "Dato",               12),
        ("E", "Falt av (skjerm)",   22),
        ("F", "Mest forvirrende",   28),
        ("G", "Tid totalt (min)",   12),
        ("H", "Fullførte alle 8 oppg?", 16),
        ("I", "Tillit 1-10",         9),
        ("J", "Hva ville betalt (kr)", 18),
        ("K", "Bruk vs arkitekt",   18),
        ("L", "Anbefale 1-10 (NPS)", 12),
        ("M", "Beste sitat",        38),
    ]
    for col, label, width in headers:
        cell = ws[f"{col}5"]
        cell.value = label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[col].width = width
    ws.row_dimensions[5].height = 32

    # 5 tester rows + extras
    profiler = [
        "1. Eldre eier vurderer kjeller",
        "2. Ny boligeier første tiltak",
        "3. Nylig gjennomgått byggesak",
        "4. Borettslag / hytteeier",
        "5. Skeptiker (ikke i bransjen)",
    ]
    for i in range(5):
        row = 6 + i
        ws[f"A{row}"] = i + 1
        ws[f"C{row}"] = profiler[i]
        ws[f"D{row}"] = date.today()
        for col in "ABCDEFGHIJKLM":
            cell = ws[f"{col}{row}"]
            cell.alignment = left if col in "BCEFKM" else center
            cell.border = border
        ws.row_dimensions[row].height = 60

    # Add 3 spare rows for re-testing
    for i in range(3):
        row = 11 + i
        ws[f"A{row}"] = ""
        ws[f"C{row}"] = "Re-test (etter fix)"
        for col in "ABCDEFGHIJKLM":
            cell = ws[f"{col}{row}"]
            cell.alignment = left if col in "BCEFKM" else center
            cell.border = border
        ws.row_dimensions[row].height = 60

    # Data validation: Tillit 1-10
    dv_score = DataValidation(type="whole", operator="between",
                              formula1=1, formula2=10,
                              error="Skriv et tall fra 1 til 10",
                              errorTitle="Ugyldig",
                              prompt="Tall fra 1 (lav) til 10 (høy)")
    dv_score.add("I6:I13")
    ws.add_data_validation(dv_score)

    dv_nps = DataValidation(type="whole", operator="between",
                            formula1=1, formula2=10,
                            prompt="1=ville aldri anbefale, 10=anbefaler aktivt")
    dv_nps.add("L6:L13")
    ws.add_data_validation(dv_nps)

    dv_yn = DataValidation(type="list", formula1='"Ja,Delvis,Nei"',
                           prompt="Ja / Delvis / Nei")
    dv_yn.add("H6:H13")
    ws.add_data_validation(dv_yn)

    dv_use = DataValidation(type="list",
                            formula1='"Ville brukt appen,Hadde ringt arkitekt,Begge"',
                            prompt="Hva ville de valgt?")
    dv_use.add("K6:K13")
    ws.add_data_validation(dv_use)

    # Color scale on trust + nps
    color_rule = ColorScaleRule(
        start_type="num", start_value=1, start_color="f5b7b7",
        mid_type="num", mid_value=5, mid_color="fef3d6",
        end_type="num", end_value=10, end_color="d8ebe1")
    ws.conditional_formatting.add("I6:I13", color_rule)
    ws.conditional_formatting.add("L6:L13", color_rule)

    ws.freeze_panes = "A6"

    # ----- Sheet 2: Per-task observations -----
    ws2 = wb.create_sheet("Per oppgave")
    ws2["A1"] = "Per-oppgave-observasjon (8 oppgaver × 5 testere)"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="0a4f3c")
    ws2.merge_cells("A1:G1")

    ws2["A2"] = ("Marker hvor hver tester nølte/falt av. "
                 "Bruk x for &quot;falt av&quot;, ? for &quot;nølte&quot;, "
                 "ok for &quot;flyt&quot;.")
    ws2["A2"].font = instruction_font
    ws2.merge_cells("A2:G2")

    task_headers = ["Oppgave", "Tester 1", "Tester 2", "Tester 3",
                    "Tester 4", "Tester 5", "Mønster / kommentar"]
    for i, h in enumerate(task_headers):
        c = ws2.cell(row=4, column=i + 1, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
    ws2.row_dimensions[4].height = 28

    tasks = [
        "1. Logg inn (BankID-mock)",
        "2. Søk egen adresse (ikke i demo)",
        "3. Velg Solhellinga 24 fra liste",
        "4. Bla gjennom eiendomsinfo + tegninger på arkiv",
        "5. Start tiltak (&quot;Hva vil du gjøre?&quot;)",
        "6a. Wizard steg 1: velg rom",
        "6b. Wizard steg 2: ny bruk",
        "6c. Wizard steg 3: tilstand",
        "6d. Wizard steg 4: bekreft + beregn",
        "7. Resultat — leser de paragrafene?",
        "8. Søknadspakke — forskjell std/premium?",
    ]
    for i, t in enumerate(tasks):
        row = 5 + i
        ws2.cell(row=row, column=1, value=t).alignment = left
        for col_i in range(2, 8):
            ws2.cell(row=row, column=col_i).border = border
            ws2.cell(row=row, column=col_i).alignment = center
        ws2.cell(row=row, column=1).border = border
        ws2.row_dimensions[row].height = 24

    ws2.column_dimensions["A"].width = 40
    for col in "BCDEFG":
        ws2.column_dimensions[col].width = 14
    ws2.column_dimensions["G"].width = 32

    dv_task = DataValidation(type="list", formula1='"ok,?,x"',
                             prompt="ok=flyt · ?=nølte · x=falt av")
    dv_task.add("B5:F15")
    ws2.add_data_validation(dv_task)

    ws2.freeze_panes = "B5"

    # ----- Sheet 3: Sammendrag -----
    ws3 = wb.create_sheet("Sammendrag")
    ws3["A1"] = "Mønster-analyse etter alle 5 tester"
    ws3["A1"].font = Font(name="Calibri", size=14, bold=True, color="0a4f3c")
    ws3.merge_cells("A1:D1")

    ws3["A3"] = "KPI"
    ws3["B3"] = "Verdi"
    ws3["C3"] = "Mål"
    ws3["D3"] = "Status / handling"
    for col in "ABCD":
        c = ws3[f"{col}3"]
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border
    ws3.row_dimensions[3].height = 26

    kpis = [
        ("Snitt-tillit (1-10)",
         "=IFERROR(AVERAGE(Resultater!I6:I10),\"\")",
         "≥ 7",
         "Hvis &lt; 6: gjør arkitekt-KS til default"),
        ("Snitt-NPS (1-10)",
         "=IFERROR(AVERAGE(Resultater!L6:L10),\"\")",
         "≥ 7",
         "Hvis &lt; 6: dårlig produkt-fit"),
        ("Median-betalingsvilje (kr)",
         "=IFERROR(MEDIAN(Resultater!J6:J10),\"\")",
         "≥ 3 000",
         "Hvis &lt; 2 000: revurder forretningsmodell"),
        ("Antall som ville brukt appen",
         "=COUNTIF(Resultater!K6:K10,\"Ville brukt appen\")",
         "≥ 3 av 5",
         "Hvis &lt; 3: identifiser blocker"),
        ("Antall som fullførte alle 8 oppgaver",
         "=COUNTIF(Resultater!H6:H10,\"Ja\")",
         "≥ 4 av 5",
         "Hvis &lt; 4: skjerm-flyten må strammes"),
        ("Antall testede",
         "=COUNTA(Resultater!A6:A10)",
         "5",
         "—"),
    ]
    for i, (k, v, m, s_) in enumerate(kpis):
        row = 4 + i
        ws3.cell(row=row, column=1, value=k).alignment = left
        ws3.cell(row=row, column=2, value=v).alignment = center
        ws3.cell(row=row, column=3, value=m).alignment = center
        ws3.cell(row=row, column=4, value=s_).alignment = left
        for col_i in range(1, 5):
            ws3.cell(row=row, column=col_i).border = border
        ws3.row_dimensions[row].height = 20

    # Beslutningsboks
    ws3["A12"] = "BESLUTNING ETTER 5 TESTER"
    ws3["A12"].fill = section_fill
    ws3["A12"].font = section_font
    ws3.merge_cells("A12:D12")
    ws3.row_dimensions[12].height = 24

    decisions = [
        ("Skal vi fortsette til Stage 1 nå?",       "Ja / Nei / Trenger fix først"),
        ("ÉN ting som må fikses før Stage 1:",      "_______________________________"),
        ("Endring i strategi (hvis noen):",         "_______________________________"),
        ("Hvilke 2 funksjoner stryker vi fra Sprint 1:", "_______________________________"),
        ("Hvem ringer vi denne uken (jur/IN):",     "_______________________________"),
    ]
    for i, (q, a) in enumerate(decisions):
        row = 13 + i
        ws3.cell(row=row, column=1, value=q).alignment = left
        ws3.cell(row=row, column=2, value=a).alignment = left
        ws3.merge_cells(start_row=row, start_column=2,
                        end_row=row, end_column=4)
        for col_i in range(1, 5):
            ws3.cell(row=row, column=col_i).border = border
        ws3.row_dimensions[row].height = 24

    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 36

    # Reorder so Resultater is first (already first since it's wb.active)
    wb.save(out)
    print(f"Wrote {out}")


# ============================================================================
# 3. ATTORNEY ATTACHMENT PDF
# ============================================================================
def build_attorney_attachment():
    out = os.path.join(HERE, "advokat_vedlegg.pdf")
    s = styles()
    doc = SimpleDocTemplate(
        out, pagesize=A4,
        leftMargin=22 * mm, rightMargin=22 * mm,
        topMargin=22 * mm, bottomMargin=18 * mm,
        title="MittBygg — Juridisk vurdering: lempninger-modul",
        author="Miad Dimensjon AS",
    )
    story = []

    story.append(Paragraph("Juridisk vurdering: regelfortolkningsmodul",
                           s["h1"]))
    story.append(Paragraph(
        "Vedlegg til bestilling om gjennomgang. Tre konkrete spørsmål "
        "knyttet til regelfortolkning, ansvarsfordeling og forsikringsbehov "
        "for forbruker-app under utvikling.",
        s["lead"]))
    story.append(HRFlowable(width="100%", color=GRAY_200, thickness=0.5,
                            spaceAfter=10))

    # Project info
    info = [
        ["Selskap:", "Miad Dimensjon AS"],
        ["Produkt:", "MittBygg — forbruker-app for byggesøknader"],
        ["Status:", "Prototype ferdig. Stage 1 (MVP) starter når juridisk klart."],
        ["Målgruppe:", "Norske husholdninger med planlagt byggetiltak (TK1-saker)"],
        ["Kontakt:", "Miad Ahmadyar · M_E_C_C@outlook.com"],
        ["Dato:", date.today().strftime("%d.%m.%Y")],
    ]
    t = Table(info, colWidths=[35 * mm, 110 * mm])
    t.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), "Helvetica", 10),
        ("FONT", (0, 0), (0, -1), "Helvetica-Bold", 10),
        ("TEXTCOLOR", (0, 0), (0, -1), GRAY_700),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, GRAY_200),
    ]))
    story.append(t)

    # 1. Hva appen gjør
    story.append(Paragraph("1. Hva appen gjør", s["h2"]))
    story.append(Paragraph(
        "MittBygg lar boligeieren skrive inn adresse, velge et byggetiltak "
        "(eksempelvis bruksendring av kjeller til soverom), og får innen 30 "
        "sekunder en strukturert vurdering med:",
        s["body"]))

    bullets = [
        "Søknadsplikt og hjemmel (PBL § 20-1, § 20-3, § 20-4 c)",
        "Tiltaksklasse (TK1/2/3) og krav til ansvarsrett",
        "TEK17-sjekk (§§ 11-13, 12-7, 13-5, 13-7, 14-3)",
        "Lempninger som kan gjøres gjeldende iht. PBL § 31-2 og DiBK "
        "veileder HO-3/2016",
        "Konkret tiltak-liste med kostnadsestimat",
        "Strukturerte søknads-dokumenter (DiBK 5153 / 5174 + "
        "nabovarsel + brannkonsept ved behov)",
    ]
    story.append(ListFlowable(
        [ListItem(Paragraph(b, s["body"])) for b in bullets],
        bulletType="bullet", leftIndent=14, bulletFontSize=10,
        spaceBefore=2, spaceAfter=8))
    story.append(Paragraph(
        "<b>Kjernen i logikken</b> er en regelmotor (Python) som er en "
        "kodifisering av vår tolkning av TEK17, SAK10 og PBL §§ 20 og 31. "
        "Det er denne tolkningen vi ønsker en juridisk gjennomgang av "
        "før appen tilbys mot betaling.",
        s["body"]))

    story.append(PageBreak())

    # 2. Lempninger-tabellen
    story.append(Paragraph("2. Lempninger som påberopes (PBL § 31-2)",
                           s["h2"]))
    story.append(Paragraph(
        "For bygg oppført før 1. juli 2010 (eldre TEK) anvender modulen "
        "følgende lempninger ved bruksendring:",
        s["body"]))

    lemp_data = [
        ["Regel", "TEK17-krav", "Lempet krav", "Hjemmel"],
        ["Takhøyde i\noppholdsrom", "2 400 mm", "2 200 mm",
         "TEK17 § 1-2,\nPBL § 31-2,\nHO-3/2016"],
        ["Dagslys /\nglassflate", "10 % av\ngulvareal", "7-10 % m/\nfunksjonell\nvurdering",
         "TEK17 § 13-7,\nPBL § 31-2,\nNS-EN 17037"],
        ["U-verdi krav", "Hele bygget i\nenergiramme",
         "Kun nye/endrede\nbygningsdeler", "PBL § 31-2,\nTEK17 § 14"],
        ["Universell\nutforming", "Tilgjengelig\nboenhet",
         "Gjelder ikke ved\nbruksendring", "PBL § 31-2"],
        ["Lyd mellom\nboenheter", "R'w ≥ 55 dB",
         "Gjelder kun ved\nflere boenheter", "TEK17 § 13-9"],
    ]
    t = Table(lemp_data, colWidths=[35 * mm, 35 * mm, 38 * mm, 38 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), GREEN_DARK),
        ("TEXTCOLOR",  (0, 0), (-1, 0), white),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 10),
        ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.3, GRAY_200),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, GRAY_50]),
    ]))
    story.append(t)
    story.append(Spacer(1, 8))

    story.append(Paragraph(
        "<b>Krav som ALLTID gjelder</b> (kommer ikke under lempningen):",
        s["h3"]))
    always = [
        "Brann (TEK17 § 11) — minimum brannskille mot annen branncelle",
        "Rømning (TEK17 § 11-13) — rømningsvindu fra soverom",
        "Våtromsforskriften — ved nytt bad",
        "Radon (TEK17 § 13-5) — tiltak hvis måling &gt; 100 Bq/m³",
        "Konstruksjonssikkerhet (TEK17 § 10) — bygget må ikke svekkes",
        "Sikkerhetsnivå (PBL § 31-3) — må ikke bli vesentlig verre",
    ]
    story.append(ListFlowable(
        [ListItem(Paragraph(a, s["body"])) for a in always],
        bulletType="bullet", leftIndent=14, bulletFontSize=10,
        spaceBefore=2, spaceAfter=8))

    story.append(PageBreak())

    # 3. Kjeller-wizardens flyt
    story.append(Paragraph("3. Eksempel: kjellerbruksendring-flyt",
                           s["h2"]))
    story.append(Paragraph(
        "For konkretisering, slik vurderer modulen et typisk tilfelle "
        "(bygg fra 1972, kjeller 2 280 mm takhøyde, ny bruk = soverom):",
        s["body"]))

    flow = [
        ["Steg", "Sjekk", "Vår fortolkning"],
        ["1", "Takhøyde 2 280 mm",
         "Aksepteres iht. PBL § 31-2 (lempet 2 200 mm) "
         "fordi byggeår &lt; 2010"],
        ["2", "Rømningsvindu",
         "Eksisterende vindu (0,8×0,6 m) er IKKE godkjent. "
         "TEK17 § 11-13 gjelder uavhengig av byggeår. "
         "Tiltak: ny vindusbrønn + større vindu."],
        ["3", "Dagslys",
         "Glassflate 1,4 % &lt; minstekrav. "
         "Lempet krav 7 % aksepteres dersom funksjonell vurdering "
         "(NS-EN 17037) dokumenteres."],
        ["4", "Radon",
         "Hvis måling ikke foreligger: pålagt tiltak "
         "(radonsperre i nytt gulv)."],
        ["5", "Drenering",
         "Forutsetning, ikke gjenstand for lempning "
         "(PBL § 31-3 sikkerhetsnivå)."],
        ["6", "Energi",
         "Kun nye/endrede bygningsdeler (innvendig isolasjon, "
         "nytt vindu) skal oppfylle TEK17 § 14-3. "
         "Eksisterende kjellervegg ikke belastet."],
        ["7", "Søknadstype",
         "PBL § 20-1 d (bruksendring). Fra tilleggsdel til "
         "hoveddel innen samme boenhet → § 20-4 c (uten ansvarsrett)."],
    ]
    t = Table(flow, colWidths=[10 * mm, 35 * mm, 105 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), GREEN_DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 9),
        ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (1, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.3, GRAY_200),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white, GRAY_50]),
    ]))
    story.append(t)
    story.append(Spacer(1, 12))

    # 4. Tre konkrete spørsmål
    story.append(Paragraph("4. Tre konkrete spørsmål til vurdering",
                           s["h2"]))

    qs = [
        ("Er fortolkningen av PBL § 31-2 og lempet takhøyde 2 200 mm "
         "holdbar uten konkret saksvurdering pr. bygg?",
         "Vi anvender lempningen automatisk når byggeår &lt; 2010. "
         "Holder dette generelt, eller må DiBKs veileder HO-3/2016 "
         "anvendes i hvert tilfelle? Hva sier nyere praksis fra "
         "kommunenes byggesaksavdelinger og statsforvalter?"),
        ("Hvilken disclaimer-konstruksjon overfører tilstrekkelig "
         "ansvar til brukeren?",
         "Forslag til ordlyd: <i>&quot;MittBygg gir veiledning basert på "
         "TEK17/SAK10/PBL. Resultatet er ikke en juridisk garanti. "
         "Endelig vedtak fattes av kommunen.&quot;</i> Holder dette? "
         "Bør den vises før eller etter regelsjekken? Bør brukeren "
         "aktivt akseptere? Behov for separat avtalevilkår?"),
        ("Trenger vi profesjonsansvarsforsikring før vi tar betalt?",
         "Vi planlegger 4 900 kr for standard søknadspakke (uten "
         "arkitekt-KS) og 12 900 kr med arkitekt-signatur. "
         "Sistnevnte involverer en autorisert arkitekt som "
         "kvalitetssikrer og signerer. "
         "I hvilke scenarier er Miad Dimensjon AS ansvarssubjekt? "
         "Når kreves ansvarsforsikring etter forsikringsavtaleloven, og "
         "hvilke forsikringer (rådgivende ingeniør / tjenesteyter / "
         "produktsansvar) er aktuelle?"),
    ]
    for i, (q, d) in enumerate(qs, 1):
        story.append(Paragraph(f"<b>Spørsmål {i}.</b> {q}", s["h3"]))
        story.append(Paragraph(d, s["body"]))
        story.append(Spacer(1, 4))

    # 5. Tilleggsmateriell
    story.append(Paragraph("5. Tilleggsmateriell tilgjengelig på forespørsel",
                           s["h2"]))
    extras = [
        "Komplett kildekode for regelfortolkningsmodulen (Python, "
        "ca. 1 500 linjer, kommentert)",
        "Demo-tilgang til prototype (privat lenke)",
        "Eksempel-rapport for kjellerbruksendring (.docx, generert)",
        "Liste over alle TEK17/SAK10/PBL-paragrafer som modulen "
        "tolker, med konkret kode-referanse",
        "Mulighet for skjermdeling-gjennomgang (45 min)",
    ]
    story.append(ListFlowable(
        [ListItem(Paragraph(e, s["body"])) for e in extras],
        bulletType="bullet", leftIndent=14, bulletFontSize=10,
        spaceBefore=2, spaceAfter=8))

    story.append(Spacer(1, 12))
    story.append(HRFlowable(width="100%", color=GRAY_200, thickness=0.5))
    story.append(Paragraph(
        "<i>Dette dokumentet inneholder forretningssensitiv informasjon "
        "og er ment kun for mottakers vurdering. Beskyttelse av "
        "konfidensialitet forventes.</i>",
        s["tiny"]))

    def cb(canv, doc_):
        header_footer(canv, doc_,
                      title="Juridisk vurdering — MittBygg",
                      subtitle=f"Miad Dimensjon AS · "
                               f"{date.today().strftime('%d.%m.%Y')}")
    doc.build(story, onFirstPage=cb, onLaterPages=cb)
    print(f"Wrote {out}")


if __name__ == "__main__":
    build_test_script()
    build_spreadsheet()
    build_attorney_attachment()
